import os
import pandas as pd
from datetime import datetime
import re
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import FORMAT_PERCENTAGE, FORMAT_NUMBER_00
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from tkcalendar import DateEntry
from typing import List, Dict, Tuple, Optional, Set, Union

class ExcelDataProcessor:
    def __init__(self):
        self.wb = None
        self.date_sheets = []
        self.team_leaders = set()
        self.master_headers = set()
        self.date_format = r'^\d{2}\.\d{2}\.\d{4}$'  # Regex for dd.mm.yyyy format
        
        # Color definitions for conditional formatting
        self.COLORS = {
            'GREEN': 'FF00B050',  # Green
            'YELLOW': 'FFFFC000', # Yellow
            'ORANGE': 'FFFF8000', # Orange
            'RED': 'FFFF0000',    # Red
            'LIGHT_GREEN': 'FFE2EFDA',  # Light green for header
            'DARK_BLUE': 'FF003366',    # Dark blue for header
        }
        
    def is_date_sheet(self, sheet_name: str) -> bool:
        """Check if sheet name matches the date format dd.mm.yyyy"""
        return bool(re.match(self.date_format, sheet_name))
    
    def parse_date(self, date_str: str) -> datetime:
        """Convert date string to datetime object"""
        return datetime.strptime(date_str, '%d.%m.%Y')
    
    def get_team_leaders(self, df: pd.DataFrame) -> Set[str]:
        """Extract unique team leaders from a DataFrame"""
        # Try to find team leader column (case insensitive)
        team_leader_cols = [col for col in df.columns if any(
            term in str(col).lower() 
            for term in ['team leader', 'team_leader', 'teamleader', 'supervisor']
        )]
        
        if not team_leader_cols:
            return set()
            
        # Use the first matching column
        team_leader_col = team_leader_cols[0]
        return set(df[team_leader_col].dropna().astype(str).unique())
    
    def process_workbook(self, file_path: str):
        """Process the Excel workbook to find date sheets and team leaders"""
        try:
            # Load the workbook
            self.wb = pd.ExcelFile(file_path)
            
            # Find all date sheets
            self.date_sheets = [sheet for sheet in self.wb.sheet_names 
                              if self.is_date_sheet(sheet)]
            
            if not self.date_sheets:
                messagebox.showerror("Error", "No sheets found with date format dd.mm.yyyy")
                return False
            
            # Sort sheets by date
            self.date_sheets.sort(key=lambda x: self.parse_date(x))
            
            # Find unique team leaders across all date sheets
            for sheet in self.date_sheets:
                df = pd.read_excel(file_path, sheet_name=sheet, header=0)
                self.team_leaders.update(self.get_team_leaders(df))
            
            if not self.team_leaders:
                messagebox.showerror("Error", "No team leaders found in any sheets")
                return False
                
            return True
            
        except Exception as e:
            messagebox.showerror("Error", f"Error processing workbook: {str(e)}")
            return False
    
    def filter_data(self, file_path: str, team_leader: str, start_date: str, end_date: str) -> pd.DataFrame:
        """Filter data based on team leader and date range"""
        try:
            # Convert input dates to datetime objects
            start_dt = self.parse_date(start_date)
            end_dt = self.parse_date(end_date)
            
            # Filter sheets within date range
            filtered_sheets = [
                sheet for sheet in self.date_sheets 
                if start_dt <= self.parse_date(sheet) <= end_dt
            ]
            
            if not filtered_sheets:
                messagebox.showwarning("No Data", "No sheets found in the specified date range")
                return pd.DataFrame()
            
            # Process each sheet and combine data
            all_data = []
            
            for sheet in filtered_sheets:
                df = pd.read_excel(file_path, sheet_name=sheet, header=0)
                
                # Find team leader column
                team_leader_cols = [col for col in df.columns if any(
                    term in str(col).lower() 
                    for term in ['team leader', 'team_leader', 'teamleader', 'supervisor']
                )]
                
                if not team_leader_cols:
                    continue
                    
                team_leader_col = team_leader_cols[0]
                
                # Filter rows for the selected team leader
                filtered = df[df[team_leader_col].astype(str).str.contains(team_leader, case=False, na=False)].copy()
                
                if not filtered.empty:
                    # Add source sheet name
                    filtered['Source_Sheet'] = sheet
                    all_data.append(filtered)
            
            if not all_data:
                messagebox.showinfo("No Data", f"No data found for team leader: {team_leader}")
                return pd.DataFrame()
                
            # Combine all data
            result = pd.concat(all_data, ignore_index=True)
            return result
            
        except Exception as e:
            messagebox.showerror("Error", f"Error filtering data: {str(e)}")
            return pd.DataFrame()
    
    def create_summary_tables(self, filtered_data: pd.DataFrame, output_path: str):
        """Create summary tables for each person with formatting"""
        if filtered_data.empty:
            return False
            
        try:
            # Create a new workbook and select the active worksheet
            wb = Workbook()
            ws = wb.active
            ws.title = "Summary"
            
            # Write headers
            headers = list(filtered_data.columns)
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color=self.COLORS['DARK_BLUE'], 
                                      end_color=self.COLORS['DARK_BLUE'], 
                                      fill_type="solid")
            
            # Write data
            for row_num, row in enumerate(filtered_data.itertuples(), 2):
                for col_num, value in enumerate(row[1:], 1):
                    ws.cell(row=row_num, column=col_num, value=value)
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save the workbook
            wb.save(output_path)
            return True
            
        except Exception as e:
            messagebox.showerror("Error", f"Error creating summary tables: {str(e)}")
            return False

class ExcelProcessorApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Excel Data Processor")
        self.root.geometry("600x500")
        
        self.processor = ExcelDataProcessor()
        self.file_path = ""
        self.output_path = ""
        
        self.setup_ui()
    
    def setup_ui(self):
        # File selection
        ttk.Label(self.root, text="Excel File:").pack(pady=(10, 0), anchor='w', padx=10)
        
        file_frame = ttk.Frame(self.root)
        file_frame.pack(fill='x', padx=10, pady=5)
        
        self.file_entry = ttk.Entry(file_frame)
        self.file_entry.pack(side='left', fill='x', expand=True)
        
        ttk.Button(file_frame, text="Browse...", command=self.browse_file).pack(side='left', padx=(5, 0))
        
        # Output location
        ttk.Label(self.root, text="Output File:").pack(pady=(10, 0), anchor='w', padx=10)
        
        output_frame = ttk.Frame(self.root)
        output_frame.pack(fill='x', padx=10, pady=5)
        
        self.output_entry = ttk.Entry(output_frame)
        self.output_entry.pack(side='left', fill='x', expand=True)
        
        ttk.Button(output_frame, text="Browse...", command=self.browse_output).pack(side='left', padx=(5, 0))
        
        # Team Leader selection
        ttk.Label(self.root, text="Team Leader:").pack(pady=(10, 0), anchor='w', padx=10)
        
        self.team_leader_var = tk.StringVar()
        self.team_leader_combo = ttk.Combobox(self.root, textvariable=self.team_leader_var, state='readonly')
        self.team_leader_combo.pack(fill='x', padx=10, pady=5)
        
        # Date range selection
        date_frame = ttk.LabelFrame(self.root, text="Date Range")
        date_frame.pack(fill='x', padx=10, pady=10, ipady=5)
        
        ttk.Label(date_frame, text="Start Date:").grid(row=0, column=0, padx=5, pady=5, sticky='w')
        self.start_date = DateEntry(date_frame, width=12, background='darkblue',
                                  foreground='white', borderwidth=2, date_pattern='dd.mm.yyyy')
        self.start_date.grid(row=0, column=1, padx=5, pady=5, sticky='w')
        
        ttk.Label(date_frame, text="End Date:").grid(row=0, column=2, padx=5, pady=5, sticky='w')
        self.end_date = DateEntry(date_frame, width=12, background='darkblue',
                                foreground='white', borderwidth=2, date_pattern='dd.mm.yyyy')
        self.end_date.grid(row=0, column=3, padx=5, pady=5, sticky='w')
        
        # Process button
        self.process_btn = ttk.Button(self.root, text="Process Data", command=self.process_data, state='disabled')
        self.process_btn.pack(pady=20)
        
        # Status label
        self.status_var = tk.StringVar()
        self.status_var.set("Select an Excel file to begin")
        ttk.Label(self.root, textvariable=self.status_var, wraplength=580).pack(fill='x', padx=10, pady=5)
    
    def browse_file(self):
        file_path = filedialog.askopenfilename(
            title="Select Excel File",
            filetypes=(("Excel files", "*.xlsx *.xls"), ("All files", "*.*"))
        )
        
        if file_path:
            self.file_path = file_path
            self.file_entry.delete(0, tk.END)
            self.file_entry.insert(0, file_path)
            
            # Set default output path
            base_name = os.path.splitext(file_path)[0]
            self.output_path = f"{base_name}_filtered.xlsx"
            self.output_entry.delete(0, tk.END)
            self.output_entry.insert(0, self.output_path)
            
            # Process the workbook
            if self.processor.process_workbook(file_path):
                # Update team leader dropdown
                self.team_leader_combo['values'] = sorted(list(self.processor.team_leaders))
                if self.processor.team_leaders:
                    self.team_leader_combo.current(0)
                
                # Set default date range to cover all sheets
                if self.processor.date_sheets:
                    start_date = self.processor.parse_date(self.processor.date_sheets[0])
                    end_date = self.processor.parse_date(self.processor.date_sheets[-1])
                    
                    # Set the date entries
                    self.start_date.set_date(start_date)
                    self.end_date.set_date(end_date)
                    
                    self.status_var.set(f"Loaded {len(self.processor.date_sheets)} date sheets with {len(self.processor.team_leaders)} team leaders")
                    self.process_btn['state'] = 'normal'
    
    def browse_output(self):
        if not self.file_path:
            messagebox.showwarning("Warning", "Please select an input file first")
            return
            
        default_name = os.path.splitext(os.path.basename(self.file_path))[0] + "_filtered.xlsx"
        output_path = filedialog.asksaveasfilename(
            title="Save Output As",
            defaultextension=".xlsx",
            initialfile=default_name,
            filetypes=(("Excel files", "*.xlsx"), ("All files", "*.*"))
        )
        
        if output_path:
            self.output_path = output_path
            self.output_entry.delete(0, tk.END)
            self.output_entry.insert(0, output_path)
    
    def process_data(self):
        if not self.file_path or not os.path.exists(self.file_path):
            messagebox.showerror("Error", "Please select a valid Excel file")
            return
            
        team_leader = self.team_leader_var.get()
        if not team_leader:
            messagebox.showerror("Error", "Please select a team leader")
            return
            
        start_date = self.start_date.get_date().strftime('%d.%m.%Y')
        end_date = self.end_date.get_date().strftime('%d.%m.%Y')
        
        # Validate date range
        try:
            start_dt = self.processor.parse_date(start_date)
            end_dt = self.processor.parse_date(end_date)
            
            if start_dt > end_dt:
                messagebox.showerror("Error", "Start date cannot be after end date")
                return
                
        except ValueError as e:
            messagebox.showerror("Error", f"Invalid date format: {str(e)}")
            return
        
        # Process the data
        self.status_var.set("Processing data, please wait...")
        self.root.update()
        
        filtered_data = self.processor.filter_data(self.file_path, team_leader, start_date, end_date)
        
        if not filtered_data.empty:
            success = self.processor.create_summary_tables(filtered_data, self.output_path)
            if success:
                self.status_var.set(f"Processing complete! Results saved to:\n{self.output_path}")
                
                # Ask if user wants to open the file
                if messagebox.askyesno("Success", "Processing complete! Would you like to open the output file?"):
                    try:
                        os.startfile(self.output_path)
                    except Exception as e:
                        messagebox.showerror("Error", f"Could not open the file: {str(e)}")
            else:
                self.status_var.set("Error occurred while creating summary tables")

def main():
    root = tk.Tk()
    app = ExcelProcessorApp(root)
    root.mainloop()

if __name__ == "__main__":
    main()
