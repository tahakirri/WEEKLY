import streamlit as st
import pandas as pd
import os
import re
from datetime import datetime, timedelta
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import tempfile
from io import BytesIO

# Set page config
st.set_page_config(
    page_title="Excel Data Processor",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Custom CSS for better styling
st.markdown("""
    <style>
    .main .block-container {
        padding-top: 2rem;
        padding-bottom: 2rem;
    }
    .stButton>button {
        width: 100%;
    }
    .stDownloadButton>button {
        width: 100%;
    }
    </style>
""", unsafe_allow_html=True)

class ExcelDataProcessor:
    def __init__(self):
        self.date_format = r'^\d{2}\.\d{2}\.\d{4}$'  # Regex for dd.mm.yyyy format
        self.COLORS = {
            'GREEN': '00B050',
            'YELLOW': 'FFC000',
            'ORANGE': 'FF8000',
            'RED': 'FF0000',
            'LIGHT_GREEN': 'E2EFDA',
            'DARK_BLUE': '003366',
        }
    
    def is_date_sheet(self, sheet_name: str) -> bool:
        """Check if sheet name matches the date format dd.mm.yyyy"""
        return bool(re.match(self.date_format, sheet_name))
    
    def parse_date(self, date_str: str) -> datetime:
        """Convert date string to datetime object"""
        return datetime.strptime(date_str, '%d.%m.%Y')
    
    def get_team_leaders(self, file_path: str) -> set:
        """Extract unique team leaders from all sheets in the workbook"""
        team_leaders = set()
        xls = pd.ExcelFile(file_path)
        
        for sheet in xls.sheet_names:
            if self.is_date_sheet(sheet):
                df = pd.read_excel(xls, sheet_name=sheet, nrows=10)  # Read only first 10 rows for headers
                team_leader_cols = [col for col in df.columns if any(
                    term in str(col).lower() 
                    for term in ['team leader', 'team_leader', 'teamleader', 'supervisor']
                )]
                
                if team_leader_cols:
                    team_leaders.update(df[team_leader_cols[0]].dropna().astype(str).unique())
        
        return team_leaders
    
    def process_workbook(self, file_path: str) -> dict:
        """Process the Excel workbook and return metadata"""
        try:
            xls = pd.ExcelFile(file_path)
            date_sheets = [sheet for sheet in xls.sheet_names if self.is_date_sheet(sheet)]
            
            if not date_sheets:
                return {"error": "No sheets found with date format dd.mm.yyyy"}
            
            # Sort sheets by date
            date_sheets.sort(key=lambda x: self.parse_date(x))
            
            # Get date range
            start_date = self.parse_date(date_sheets[0])
            end_date = self.parse_date(date_sheets[-1])
            
            # Get team leaders
            team_leaders = sorted(list(self.get_team_leaders(file_path)))
            
            return {
                "success": True,
                "date_sheets": date_sheets,
                "start_date": start_date,
                "end_date": end_date,
                "team_leaders": team_leaders
            }
            
        except Exception as e:
            return {"error": f"Error processing workbook: {str(e)}"}
    
    def filter_data(self, file_path: str, team_leader: str, start_date: str, end_date: str) -> pd.DataFrame:
        """Filter data based on team leader and date range"""
        try:
            xls = pd.ExcelFile(file_path)
            all_data = []
            
            for sheet in xls.sheet_names:
                if not self.is_date_sheet(sheet):
                    continue
                    
                sheet_date = self.parse_date(sheet)
                start_dt = self.parse_date(start_date)
                end_dt = self.parse_date(end_date)
                
                if start_dt <= sheet_date <= end_dt:
                    df = pd.read_excel(xls, sheet_name=sheet)
                    
                    # Find team leader column
                    team_leader_cols = [col for col in df.columns if any(
                        term in str(col).lower() 
                        for term in ['team leader', 'team_leader', 'teamleader', 'supervisor']
                    )]
                    
                    if team_leader_cols:
                        team_leader_col = team_leader_cols[0]
                        filtered = df[df[team_leader_col].astype(str).str.contains(team_leader, case=False, na=False)].copy()
                        
                        if not filtered.empty:
                            filtered['Source_Sheet'] = sheet
                            all_data.append(filtered)
            
            if not all_data:
                return pd.DataFrame()
                
            return pd.concat(all_data, ignore_index=True)
            
        except Exception as e:
            st.error(f"Error filtering data: {str(e)}")
            return pd.DataFrame()

def create_excel_download_link(df: pd.DataFrame, filename: str = "filtered_data.xlsx") -> str:
    """Generate a download link for the filtered data"""
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Filtered_Data')
        
        # Get the workbook and worksheet objects
        workbook = writer.book
        worksheet = writer.sheets['Filtered_Data']
        
        # Format headers
        for col_num, value in enumerate(df.columns.values):
            worksheet.cell(row=1, column=col_num+1).font = Font(bold=True, color="FFFFFF")
            worksheet.cell(row=1, column=col_num+1).fill = PatternFill(
                start_color="003366", end_color="003366", fill_type="solid"
            )
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    # Create download link
    b64 = output.getvalue()
    return b64

def main():
    st.title("📊 Excel Data Processor")
    st.markdown("Process Excel files with date-formatted sheets and filter by team leader and date range.")
    
    # Initialize session state
    if 'processed_data' not in st.session_state:
        st.session_state.processed_data = None
    
    # File upload
    uploaded_file = st.file_uploader("Upload Excel File", type=['xlsx', 'xls'])
    
    if uploaded_file is not None:
        # Save uploaded file to a temporary file
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
            tmp_file.write(uploaded_file.getvalue())
            tmp_file_path = tmp_file.name
        
        try:
            # Initialize processor
            processor = ExcelDataProcessor()
            
            # Process workbook
            result = processor.process_workbook(tmp_file_path)
            
            if "error" in result:
                st.error(result["error"])
            else:
                # Date range selector
                col1, col2 = st.columns(2)
                with col1:
                    start_date = st.date_input(
                        "Start Date",
                        value=result["start_date"],
                        min_value=result["start_date"],
                        max_value=result["end_date"]
                    )
                with col2:
                    end_date = st.date_input(
                        "End Date",
                        value=result["end_date"],
                        min_value=result["start_date"],
                        max_value=result["end_date"]
                    )
                
                # Team leader selection
                team_leader = st.selectbox(
                    "Select Team Leader",
                    options=result["team_leaders"]
                )
                
                # Process button
                if st.button("Process Data", type="primary"):
                    with st.spinner("Processing data..."):
                        # Filter data
                        filtered_data = processor.filter_data(
                            tmp_file_path,
                            team_leader,
                            start_date.strftime('%d.%m.%Y'),
                            end_date.strftime('%d.%m.%Y')
                        )
                        
                        if not filtered_data.empty:
                            st.session_state.processed_data = filtered_data
                            st.success(f"Found {len(filtered_data)} records matching your criteria!")
                            
                            # Show preview
                            st.subheader("Preview of Filtered Data")
                            st.dataframe(filtered_data.head(), use_container_width=True)
                            
                            # Download button
                            st.download_button(
                                label="Download Filtered Data",
                                data=create_excel_download_link(filtered_data),
                                file_name=f"filtered_data_{team_leader.replace(' ', '_')}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )
                        else:
                            st.warning("No data found matching your criteria.")
            
        except Exception as e:
            st.error(f"An error occurred: {str(e)}")
        finally:
            # Clean up temporary file
            try:
                os.unlink(tmp_file_path)
            except:
                pass
    
    # Add some space at the bottom
    st.markdown("---")
    st.markdown("### How to Use")
    st.markdown("""
    1. Upload your Excel file with date-formatted sheets (dd.mm.yyyy)
    2. Select the date range you want to analyze
    3. Choose a team leader from the dropdown
    4. Click 'Process Data' to filter the data
    5. Download the filtered results
    """)

if __name__ == "__main__":
    main()
